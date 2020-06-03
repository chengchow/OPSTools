# -*- coding: utf-8 -*-
"""

"""
## 导入模块
from __future__ import division
from __future__ import print_function
import os
import sys
import logging
import logging.config
#import numpy as np
import pandas as pd
#import subprocess
import itertools
#import psutil
#import string
import argparse
from openpyxl import load_workbook
from configparser import ConfigParser as _conf

## 获取根目录路径
homePath = nowPath = os.path.dirname(os.path.abspath(__file__))

## 配置日志格式
logging.config.fileConfig(os.path.join(homePath, 'conf/logging.ini'))

## 读取配置文件
conFile = os.path.join(homePath, 'conf/config.ini')
conf = _conf(allow_no_value=True)
conf.read(conFile)

## 设置应用文件路径
appPath = os.path.join(homePath, 'app')

## 添加app路径到环境变量中
sys.path.append(appPath)

## 附件存在路径
tmpPath = os.path.join(homePath, 'tmp')

## sheet名称
sheets = ('压测执行时间', '平均请求时间')

## 定义参数 
def args():
    parser=argparse.ArgumentParser(description=__doc__)
    parser.add_argument('-l', '--count_list', action='store', dest='count_list',
        help="压测次数列表, 多个以','分隔. "
    )
    ## 压测类型为插入, 修改还是查询
    parser.add_argument(
        '-t', '--type', action='store', dest='_type', type=int, default=0,
        help="压测类型(0|1|2), 依次插入|修改|查询. 默认0"
    )
    ## 压测结束是否删除压测数据库
    parser.add_argument(
        '-d', '--delete', action='store_true', default=False, dest='deleted', 
        help="删除压测数据库"
    )
    ## 插入数据前是否初始数据库
    parser.add_argument(
        '-i', '--init', action='store_true', default=False, dest='init',
        help="初始化数据库, 结合参数-t0使用"
    )

    return parser.parse_args()

## 获取压测每个并发数据
def get_stress_list(total, threadsList, stressDict):
    from mysqlstress import main as mysqlstress

    re =[ mysqlstress(**dict(**stressDict, **dict(row=x, col=total//x))) for x in threadsList ]
    return re

## numpy数组添加菜单写入excel文件
def numpy_to_excel(numpyDict, xlsName, rowIndex, colIndex):
    writer = pd.ExcelWriter(xlsName)

    for key,value in numpyDict.items():
        data = pd.DataFrame(value)
    
        data.columns = colIndex
        data.index = rowIndex

        data.to_excel(writer, key)
        writer.save()

    writer.close()

## 修改存在Exce首行首列数据
def excel_modify(xlsName, value):
    wb = load_workbook(filename=xlsName)

    for x in wb.sheetnames:
        sheetName = wb[x]
        sheetName['A1'] = value

    wb.save(xlsName)
    wb.close()

## 发送邮件
def send_mail(mailConn):
    from sendemail import main as sendemail
    sendemail(**mailConn)

## 主程序    
def main():
    ## 获取参数
    parser = args()
    
    ## 获取压测次数列表, 优先去命令参数, 其次去配置参数
    if parser.count_list:
        countList = [ int(x) for x in parser.count_list.replace(' ','').split(',') ]
    else:
        countList = [ int(x) for x in conf.get('general', 'stress_count').replace(' ','').split(',') ]
    
    ## 并发列表
    threadsList = [ int(x) for x in conf.get('general', 'threads_list').replace(' ','').split(',') ]

    ## 获取存储文件名称
    xlsName = conf.get('mail', 'att_file')
    xlsFile = os.path.join(tmpPath, xlsName)

    ## 配置连接字典
    connDict = dict(
        host = conf.get('db', 'db_host'),
        user = conf.get('db', 'db_user'),
        password = conf.get('db', 'db_pass'),
        port = int(conf.get('db', 'db_port')),
        _type = parser._type,
        _del = parser.deleted,
        _init = parser.init
    )

    ## 压测执行时间汇总
    totalTimeList = [ get_stress_list(x, threadsList, connDict) for x in countList ]

    ## 平均请求时间汇总
    ride = itertools.cycle(threadsList)
    divList = countList*len(threadsList)
    divList.sort()
    div = itertools.cycle(divList)
    try:
        perTimeList = [ list(map(lambda y:round(y*next(ride)/next(div), 4), x)) for x in totalTimeList ] 
    except TypeError:
        logging.error('获取数据为空{}'.format(totalTimeList))
    else:
        ## 转成numpy_to_execl识别的字典
        numpyDict = dict(zip(sheets, (totalTimeList, perTimeList)))

        ## 写入excel
        numpy_to_excel(numpyDict, xlsFile, countList, threadsList)

        ## 修改execl文件
        excel_modify(xlsFile, '总数|并发')

    if os.path.isfile(xlsFile):
        ## 发送邮件信息
        mailConn = dict(
            toMail = conf.get('mail', 'to_mail'),
            sub = conf.get('mail', 'subject'),
            context = conf.get('mail', 'context'),
            xlsFile = xlsFile
        )
        ## 发送邮件
        send_mail(mailConn)
    else:
        logging.error('压测结果文件不存在')


## 执行脚本
if __name__ == '__main__':
    main()
